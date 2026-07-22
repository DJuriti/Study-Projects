from openpyxl import Workbook, load_workbook
from pathlib import Path

num = 0
arquivo = Path('teste.xlsx')

if arquivo.exists():
    planilha = load_workbook('teste.xlsx')
    aba = planilha.active
else:
    planilha = Workbook()
    aba = planilha.active
    aba.title = 'Estoque'
    aba.append(['Produto', 'Quantidade', 'Preço'])

print('=' * 35)
print("             ESTOQUE")
print('=' * 35)

while True:
    if num == 7:
        break
    else:
        print('1 - CADASTRAR PRODUTO')
        print('2 - CONSULTAR PRODUTO')
        print('3 - ATUALIZAR QUANTIDADE')
        print('4 - ATUALIZAR PREÇO')
        print('5 - LISTAR PRODUTOS')
        print('6 - REMOVER PRODUTO')
        print('7 - SAIR')
        print('=' * 35)

        num = int(input('Sua opção: '))
        print('=' * 35)

        # 1 - CADASTRAR
        if num == 1:
            produto = str(input('Nome do produto: '))
            quantidade = int(input('Quantidade: '))
            preco = float(input('Preço: '))
            aba.append([produto, quantidade, preco])
            print('Produto adicionado com sucesso!')
            print(f'Total de linhas: {aba.max_row}')
            print('=' * 35)

        # 2 - CONSULTAR
        elif num == 2:
            produto = str(input('Nome do produto: '))
            encontrou = False
            for linha in aba.iter_rows(min_row=2):
                if linha[0].value == produto:
                    print(f'Produto: {linha[0].value}')
                    print(f'Quantidade: {linha[1].value}')
                    print(f'Preço: {linha[2].value}')
                    encontrou = True
                    break

            if not encontrou:
                print("Produto não encontrado.")
            print('=' * 35)

        # 3 - ATUALIZAR QUANTIDADE
        elif num == 3:
            produto = str(input('Nome do produto: '))
            encontrou = False

            for linha in aba.iter_rows(min_row=2):
                if linha[0].value == produto:
                    nova_quantidade = int(input('Nova quantidade: '))
                    linha[1].value = nova_quantidade
                    encontrou = True
                    print('Quantidade atualizada com sucesso!')
                    break

            if not encontrou:
                print('Produto não encontrado.')
            print('=' * 35)

        # 4 - ATUALIZAR PREÇO
        elif num == 4:
            produto = str(input('Nome do produto: '))
            encontrou = False

            for linha in aba.iter_rows(min_row=2):
                if linha[0].value == produto:
                    novo_preco = float(input('Novo preço: '))
                    linha[2].value = novo_preco
                    encontrou = True
                    print('Preço atualizado com sucesso!')
                    break

            if not encontrou:
                print('Produto não encontrado.')
            print('=' * 35)

        # 5 - LISTAR PRODUTOS
        elif num == 5:
            for linha in aba.iter_rows(min_row=2):
                print([celula.value for celula in linha])
            print('=' * 35)

        # 6 - REMOVER PRODUTO
        elif num == 6:
            produto = str(input('Nome do produto: '))
            encontrou = False

            for linha in aba.iter_rows(min_row=2):
                if linha[0].value == produto:
                    aba.delete_rows(linha[0].row)
                    encontrou = True
                    print('Produto removido com sucesso!')
                    break

            if not encontrou:
                print('Produto não encontrado.')
            print('=' * 35)

        # 7 - SAIR
        elif num == 7:
            break

        else:
            print('Opção inválida!')

print('Salvando planilha...')
planilha.save('teste.xlsx')
print('Planilha salva com sucesso!')