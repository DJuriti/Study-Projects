from openpyxl import load_workbook
from pathlib import Path
from reportlab.pdfgen import canvas

# Mostra o caminho absoluto da pasta onde o programa está sendo executado.
# Útil para verificar onde o Python está procurando e salvando arquivos.
print(Path().resolve())

# Cria um novo arquivo PDF chamado "relatorio.pdf".
pdf = canvas.Canvas('relatorio.pdf')

# Abre a planilha do Excel que contém os produtos.
planilha = load_workbook('teste.xlsx')

# Seleciona a aba ativa da planilha.
aba = planilha.active

# Define a posição inicial no eixo Y onde os produtos serão escritos.
y = 730

# Escreve o título do relatório no topo da página.
pdf.drawString(100, 750, 'RELATÓRIO DE ESTOQUE')

# Percorre todas as linhas da planilha, ignorando o cabeçalho.
for linha in aba.iter_rows(min_row=2):

    # Escreve o nome do produto.
    pdf.drawString(100, y, f'Produto: {linha[0].value}')

    # Escreve a quantidade do produto.
    pdf.drawString(100, y - 20, f'Quantidade: {linha[1].value}')

    # Escreve o preço do produto.
    pdf.drawString(100, y - 40, f'Preço: R$ {linha[2].value}')

    # Move a posição para baixo para escrever o próximo produto.
    y -= 60

# Salva e finaliza o arquivo PDF.
pdf.save()