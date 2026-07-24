import pyautogui
import time
from openpyxl import load_workbook
import subprocess

planilha = load_workbook('testerobo.xlsx')
aba = planilha.active

subprocess.Popen('notepad.exe') #Executa o Bloco de Notas
time.sleep(5)

for linha in aba.iter_rows(min_row=2):
    produto = linha[0].value
    quantidade = linha[1].value
    preco = linha[2].value

    texto = (
        f"Produto: {produto}\n"
        f"Quantidade: {quantidade}\n"
        f"Preco: {preco}\n\n"
    )

    pyautogui.write(texto, interval=0.05)

pyautogui.hotkey('ctrl', 's')
time.sleep(2)
pyautogui.write("relatorio.txt")
time.sleep(2)
pyautogui.press('enter')
time.sleep(2)
pyautogui.hotkey('alt', 'f4')